import os
import openpyxl
import numpy as np

def groundwater_model_unconfined(excel_file):
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    # Read simulation variables from "Main" sheet
    main_sheet = wb["Main"]
    numx = main_sheet['B6'].value
    numy = main_sheet['B7'].value
    cell_size = main_sheet['B9'].value
    time_step = main_sheet['B10'].value
    time_total = main_sheet['B15'].value
    obs_x = main_sheet['B19'].value - 1  # Adjust for 0-based indexing
    obs_y = main_sheet['C19'].value - 1  # Adjust for 0-based indexing
    
    # Initialize arrays
    HC = np.zeros((numx, numy))
    Sy = np.zeros((numx, numy))
    gselev = np.zeros((numx, numy))
    brelev = np.zeros((numx, numy))
    rech = np.zeros((numx, numy))
    pump = np.zeros((numx, numy))
    riv_stage = np.zeros((numx, numy))
    riv_length = np.zeros((numx, numy))
    riv_width = np.zeros((numx, numy))
    riv_K = np.zeros((numx, numy))
    riv_thick = np.zeros((numx, numy))
    riv_bot = np.zeros((numx, numy))
    ibound = np.zeros((numx, numy), dtype=int)
    init_head = np.zeros((numx, numy))
    vol = np.zeros((numx, numy))
    vol_new = np.zeros((numx, numy))
    head = np.zeros((numx, numy))
    head_new = np.zeros((numx, numy))
    
    # Helper function to read 2D array from sheet
    def read_array(sheet, start_row, start_col, rows, cols):
        array = np.zeros((rows, cols))
        for j in range(rows):
            for i in range(cols):
                cell = sheet.cell(row=start_row + j, column=start_col + i)
                array[i, j] = cell.value if cell.value is not None else 0
        return array
    
    # Read aquifer properties
    HC = read_array(wb["K"], 2, 2, numy, numx)
    Sy = read_array(wb["Sy"], 2, 2, numy, numx)
    gselev = read_array(wb["GSElev"], 2, 2, numy, numx)
    brelev = read_array(wb["BRElev"], 2, 2, numy, numx)
    
    # Read sources and sinks
    rech = read_array(wb["Recharge"], 2, 2, numy, numx)
    pump = read_array(wb["Pumping"], 2, 2, numy, numx)
    riv_stage = read_array(wb["Streams"], 2, 2, numy, numx)
    riv_length = read_array(wb["Streams"], 44, 2, numy, numx)
    riv_width = read_array(wb["Streams"], 86, 2, numy, numx)
    riv_K = read_array(wb["Streams"], 128, 2, numy, numx)
    riv_thick = read_array(wb["Streams"], 170, 2, numy, numx)
    riv_bot = read_array(wb["Streams"], 212, 2, numy, numx)
    
    # Read initial conditions and boundary conditions
    init_head = read_array(wb["ICs"], 2, 2, numy, numx)
    ibound = read_array(wb["cell_type"], 2, 2, numy, numx).astype(int)
    
    # Solve for groundwater volume and head
    numts = int(time_total / time_step)  # Number of time steps
    vol_total = np.zeros(numts)
    rech_total = np.zeros(numts)
    pump_total = np.zeros(numts)
    exchange_total = np.zeros(numts)
    obs_head = np.zeros(numts)
    
    # Initialize head and volume arrays
    head = init_head.copy()
    vol = init_head * cell_size * cell_size * Sy
    
    # Loop through time steps
    current_time = 0
    for n in range(numts):
        current_time += time_step
        
        # Loop through cells (excluding boundaries)
        write_row = 3
        write_row_vol = 45
        for j in range(1, numy - 1):
            write_col = 7
            for i in range(1, numx - 1):
                # Cell area
                area = cell_size * cell_size  # m2
                
                # Calculate conductance for each face
                c_west = ((HC[i - 1, j] + HC[i, j]) / 2) * cell_size * ((head[i - 1, j] + head[i, j]) / 2) / cell_size
                c_east = ((HC[i + 1, j] + HC[i, j]) / 2) * cell_size * ((head[i + 1, j] + head[i, j]) / 2) / cell_size
                c_north = ((HC[i, j - 1] + HC[i, j]) / 2) * cell_size * ((head[i, j - 1] + head[i, j]) / 2) / cell_size
                c_south = ((HC[i, j + 1] + HC[i, j]) / 2) * cell_size * ((head[i, j + 1] + head[i, j]) / 2) / cell_size
                
                # Get head values, accounting for boundary conditions
                h_west = init_head[i - 1, j] if ibound[i - 1, j] == 2 else head[i - 1, j]
                h_east = init_head[i + 1, j] if ibound[i + 1, j] == 2 else head[i + 1, j]
                h_north = init_head[i, j - 1] if ibound[i, j - 1] == 2 else head[i, j - 1]
                h_south = init_head[i, j + 1] if ibound[i, j + 1] == 2 else head[i, j + 1]
                
                # Calculate flow rates across each face
                Q_west = c_west * (h_west - head[i, j])
                Q_east = c_east * (h_east - head[i, j])
                Q_north = c_north * (h_north - head[i, j])
                Q_south = c_south * (h_south - head[i, j])
                
                # Calculate recharge volume
                Q_recharge = (rech[i, j] / 1000) * cell_size * cell_size
                
                # Calculate aquifer-stream exchange
                Q_exchange = 0
                if riv_K[i, j] > 0:  # Check if cell has a stream
                    flow_area = riv_width[i, j] * riv_length[i, j]
                    if head[i, j] > riv_stage[i, j]:
                        head_gradient = (head[i, j] - riv_stage[i, j]) / riv_thick[i, j] * (-1)
                    elif head[i, j] < riv_stage[i, j] and head[i, j] > riv_bot[i, j]:
                        head_gradient = (riv_stage[i, j] - head[i, j]) / riv_thick[i, j]
                    elif head[i, j] < riv_bot[i, j]:
                        head_gradient = (riv_stage[i, j] - riv_bot[i, j]) / riv_thick[i, j]
                    Q_exchange = flow_area * riv_K[i, j] * head_gradient
                
                # Calculate change in groundwater storage and update head
                vol_change = (Q_west + Q_east + Q_north + Q_south + Q_recharge + pump[i, j] + Q_exchange) * time_step
                vol_new[i, j] = vol[i, j] + vol_change
                head_new[i, j] = (vol_new[i, j] / (area * Sy[i, j])) + brelev[i, j]
                
                # Write new values to "Main" sheet
                main_sheet.cell(row=write_row_vol, column=write_col).value = vol_new[i, j]
                main_sheet.cell(row=write_row, column=write_col).value = head_new[i, j]
                write_col += 1
                
                # Sum up total volumes and fluxes
                vol_total[n] += vol_new[i, j]
                rech_total[n] += Q_recharge
                pump_total[n] += pump[i, j]
                exchange_total[n] += Q_exchange
            
            write_row += 1
            write_row_vol += 1
        
        # Update current time in "Main" sheet
        main_sheet['B17'].value = current_time
        
        # Store observation cell head value
        obs_head[n] = head[obs_x, obs_y]
        
        # Update head and volume arrays
        for j in range(1, numy):
            for i in range(1, numx):
                vol[i, j] = vol_new[i, j]
                head[i, j] = head_new[i, j]
    
    # Write total volumes and fluxes to "Totals" sheet
    totals_sheet = wb["test"]
    row = 5
    current_time = 0
    for n in range(numts):
        current_time += time_step
        totals_sheet.cell(row=row, column=3).value = current_time
        totals_sheet.cell(row=row, column=4).value = vol_total[n]
        totals_sheet.cell(row=row, column=5).value = rech_total[n]
        totals_sheet.cell(row=row, column=6).value = pump_total[n]
        totals_sheet.cell(row=row, column=7).value = exchange_total[n]
        totals_sheet.cell(row=row, column=8).value = obs_head[n]
        row += 1
    
    # Save the workbook
    wb.save(excel_file + ".out.xlsx")

if __name__ == "__main__":
    wd = "D:\\Workshops\\20250623_SWAT_jeju"
    excel_file = "Model 2D Structured Grid.xlsm"
    groundwater_model_unconfined(os.path.join(wd, excel_file))